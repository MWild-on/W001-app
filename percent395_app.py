# percent395_app.py

import io
import os
import zipfile
import math
import datetime as dt
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import pandas as pd
import openpyxl
import streamlit as st

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import ParagraphStyle


# =========================
# Core: percent395_app
# =========================

@dataclass
class RateRow:
    start: dt.date
    end: dt.date
    days_in_year: int
    rate: float  # decimal: 0.16 == 16%


def run():
    percent395_app()

def _register_cyrillic_font():
    """
    Как в вашем примере: best effort.
    Возвращает (regular_font, bold_font).
    """
    try:
        pdfmetrics.registerFont(TTFont("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"))
        return "DejaVuSans", "DejaVuSans-Bold"
    except Exception:
        return "Helvetica", "Helvetica-Bold"





def percent395_app():
    st.title("Начисление процентов по ст. 395 ГК РФ")

    uploaded = st.file_uploader("Загрузка файла (Excel)", type=["xlsx"])

    col1, col2 = st.columns(2)
    with col1:
        date_from = st.date_input("Дата от (дата начала расчета)")
    with col2:
        date_to = st.date_input("Дата до (дата до которой производится расчет)")

    # кнопка расчёта
    calc = st.button("Рассчитать", type="primary", disabled=(uploaded is None))

    # --- сигнатура входа, чтобы понимать: те же ли параметры ---
    sig = None
    if uploaded is not None:
        # uploaded.name + даты достаточно. Если нужно жёстче — можно добавить size/хеш.
        sig = (uploaded.name, str(date_from), str(date_to))

    # если параметры изменились — сбрасываем старый результат
    if sig is not None and st.session_state.get("p395_sig") != sig:
        st.session_state["p395_sig"] = sig
        st.session_state.pop("p395_zip", None)
        st.session_state.pop("p395_xlsx", None)

    # валидация дат (без return до отображения кнопок скачивания)
    if uploaded is not None and date_from > date_to:
        st.error("Дата от не может быть больше даты до.")
        return

    # --- если нажали рассчитать, делаем расчёт и кладём в session_state ---
    if calc:
        if uploaded is None:
            st.error("Загрузите Excel файл.")
        else:
            try:
                zip_bytes, xlsx_bytes = run_calculation(uploaded.getvalue(), date_from, date_to)
                st.session_state["p395_zip"] = zip_bytes
                st.session_state["p395_xlsx"] = xlsx_bytes
                st.success("Готово. Скачайте результат.")


    # --- показываем кнопки скачивания ВСЕГДА, если результат уже есть ---
    if "p395_zip" in st.session_state and "p395_xlsx" in st.session_state:
        st.info("Результаты готовы. Можно скачать файлы.")
        c1, c2 = st.columns(2)

        with c1:
            st.download_button(
                "Скачать ZIP (только PDF)",
                data=st.session_state["p395_zip"],
                file_name="percent395_pdfs.zip",
                mime="application/zip",
                use_container_width=True,
                key="download_zip_pdfs",
            )

        with c2:
            st.download_button(
                "Скачать Excel",
                data=st.session_state["p395_xlsx"],
                file_name="percent395_result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_excel_395",
            )

    
    except Exception as e:
        st.exception(e)



# =========================
# Helpers
# =========================

def _to_date(x) -> Optional[dt.date]:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    return pd.to_datetime(x, dayfirst=True).date()


def _sheet_to_df(wb: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
    ws = wb[sheet_name]
    data = list(ws.values)

    header_idx = None
    for i, row in enumerate(data):
        if any(v is not None and str(v).strip() != "" for v in row):
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame()

    header = [str(v).strip() if v is not None else "" for v in data[header_idx]]
    rows = data[header_idx + 1:]
    df = pd.DataFrame(rows, columns=header).dropna(axis=1, how="all").dropna(how="all")
    return df


def _normalize_rate(val) -> float:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return 0.0
    v = float(val)
    return v / 100.0 if v > 1 else v


def _find_sheet_name(wb: openpyxl.Workbook, candidates: List[str]) -> Optional[str]:
    lower_map = {name.lower(): name for name in wb.sheetnames}
    for c in candidates:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    return None


def _parse_rates(df_rate: pd.DataFrame) -> List[RateRow]:
    col_s = next(c for c in df_rate.columns if c.strip().lower() == "с")
    col_po = next(c for c in df_rate.columns if c.strip().lower() == "по")
    col_rate = next(c for c in df_rate.columns if "ставк" in c.lower())
    col_diy = next(
        c for c in df_rate.columns
        if "дней в году" in c.lower() or ("дней" in c.lower() and "год" in c.lower())
    )

    out: List[RateRow] = []
    for _, r in df_rate.iterrows():
        s = _to_date(r[col_s])
        e = _to_date(r[col_po])
        if not s or not e:
            continue
        out.append(
            RateRow(
                start=s,
                end=e,
                days_in_year=int(r[col_diy]),
                rate=_normalize_rate(r[col_rate]),
            )
        )
    out.sort(key=lambda x: x.start)
    return out


def _rate_for_date(rates: List[RateRow], d: dt.date) -> RateRow:
    for rr in rates:
        if rr.start <= d <= rr.end:
            return rr
    raise ValueError(f"Нет ставки для даты {d}")


def _fmt_date(d: dt.date) -> str:
    return d.strftime("%d.%m.%Y")


def _fmt_money(x: float) -> str:
    return f"{x:,.2f}".replace(",", " ").replace(".", ",")


def _ensure_cyrillic_font() -> str:

    candidates = [
        ("LiberationSerif", "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf"),
        ("LiberationSerif", "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf"),

        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        ("DejaVuSans", "/usr/share/fonts/dejavu/DejaVuSans.ttf"),

        ("FreeSans", "/usr/share/fonts/truetype/freefont/FreeSans.ttf"),
    ]

    for font_name, path in candidates:
        if os.path.exists(path):
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name

    raise FileNotFoundError(
        "Не найден шрифт с кириллицей. "
        "Положите TTF в папку проекта (например fonts/DejaVuSans.ttf) "
        "и добавьте его в список candidates."
    )


def _compute_contract(
    date_from: dt.date,
    date_to: dt.date,
    principal: float,
    rates: List[RateRow],
    payments: List[Tuple[dt.date, float]],
) -> Tuple[float, List[dict]]:
    # платеж уменьшает ОД со следующего дня (payment_date + 1)
    payments = sorted([(d, a) for d, a in payments if d is not None], key=lambda x: x[0])

    pay_map: Dict[dt.date, float] = {}
    for d, a in payments:
        if date_from <= d <= date_to:
            pay_map[d] = pay_map.get(d, 0.0) + float(a)

    rate_starts = [r.start for r in rates if date_from < r.start <= date_to]
    pay_effective = [(d + dt.timedelta(days=1)) for d in pay_map.keys() if d < date_to]

    breakpoints = sorted(set([date_from, date_to + dt.timedelta(days=1)] + rate_starts + pay_effective))

    rows: List[dict] = []
    cur_principal = float(principal)
    total = 0.0

    for i in range(len(breakpoints) - 1):
        seg_start = breakpoints[i]
        seg_end = breakpoints[i + 1] - dt.timedelta(days=1)

        seg_start = max(seg_start, date_from)
        seg_end = min(seg_end, date_to)
        if seg_start > seg_end:
            continue

        rr = _rate_for_date(rates, seg_start)
        days = (seg_end - seg_start).days + 1  # inclusive
        interest = cur_principal * days * rr.rate / rr.days_in_year
        total += interest

        rows.append(
            {
                "kind": "interest",
                "from": seg_start,
                "to": seg_end,
                "days": days,
                "rate": rr.rate,
                "diy": rr.days_in_year,
                "principal": cur_principal,
                "formula": f"{cur_principal:.2f}*{days}*1/{rr.days_in_year}*{rr.rate*100:.2f}%",
                "interest": interest,
            }
        )

        # строки платежей (в пределах периода), ОД уменьшаем на дату платежа
        for pdate in sorted([d for d in pay_map.keys() if seg_start <= d <= seg_end]):
            amount = pay_map[pdate]
            cur_principal = max(0.0, cur_principal - amount)
            rows.append(
                {
                    "kind": "payment",
                    "payment_date": pdate,
                    "payment_amount": amount,
                    "principal_after": cur_principal,
                }
            )

    return total, rows


def _build_pdf(
    contract_no: str,
    contract_date: Optional[dt.date],
    fio: str,
    date_to: dt.date,
    principal: float,
    rows: List[dict],
    total: float,
) -> bytes:
    font, font_bold = _register_cyrillic_font()
    styles = getSampleStyleSheet()

    base = ParagraphStyle(
        "base",
        parent=styles["Normal"],
        fontName=font,
        fontSize=10,
        leading=12,
    )

    title_style = ParagraphStyle(
        "title",
        parent=base,
        fontName=font_bold,
        fontSize=11,
        leading=14,
        alignment=1,      # CENTER
        spaceAfter=6,
    )

    buf = io.BytesIO()
    from reportlab.lib.pagesizes import A4, landscape
    
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"395_{contract_no}",
    )

    cd = _fmt_date(contract_date) if contract_date else ""
    fio_txt = fio or ""

    header_text = (
        "Расчет процентов за неправомерное пользование чужими денежными средствами "
        "по периодам действия ключевой ставки ЦБ РФ по номеру договора "
        f"№{contract_no}"
        + (f" от {cd}" if cd else "")
        + (f" {fio_txt}" if fio_txt else "")
        + f" на {_fmt_date(date_to)} г."
    )

    elements = [
        Paragraph(header_text, title_style),
        Spacer(1, 4 * mm),
    ]

    # Шапка таблицы — как в примере (переносы через \n)
    data = [[
        "№",
        "Период\nпросрочки c",
        "Период\nпросрочки по",
        "Коли\nчеств\nо\nдней",
        "Ставка в\n%",
        "Сумма\nплатежа",
        "Дата\nплатежа",
        "Основной\nдолг",
        "Формула",
        "Сумма\nпроцентов",
    ]]

    n = 0
    for it in rows:
        if it["kind"] == "interest":
            n += 1
            pr = float(it["principal"])
            days = int(it["days"])
            rate = float(it["rate"])
            diy = int(it["diy"])
            interest = float(it["interest"])

            formula = f"{pr:.2f}*{days}*1/{diy}*{rate*100:.2f}%"
            formula = formula.replace(".", ",")

            data.append([
                str(n),
                _fmt_date(it["from"]),
                _fmt_date(it["to"]),
                str(days),
                f"{rate*100:.2f}%".replace(".", ","),
                "-",
                "-",
                _fmt_money(pr),
                formula,
                _fmt_money(round(interest + 1e-9, 2)),
            ])

        else:
            # строка платежа (как отдельная строка, без номера)
            pdate = it["payment_date"]
            pay = float(it["payment_amount"])
            pr_after = float(it["principal_after"])

            data.append([
                "",
                _fmt_date(pdate),
                "",
                "",
                "",
                _fmt_money(pay),
                _fmt_date(pdate),
                _fmt_money(pr_after),
                "",
                "",
            ])

    data.append(["", "", "", "", "", "", "", "", "Итого:", _fmt_money(total)])

    # Ширины колонок — долями от doc.width (как в примере)
    _w = [6, 14, 14, 9, 12, 14, 13, 14, 32, 14]
    _sumw = sum(_w)
    col_widths = [doc.width * (w / _sumw) for w in _w]

    tbl = Table(
        data,
        colWidths=col_widths,
        repeatRows=1,
    )

    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), font_bold),
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("GRID", (0, 0), (-1, -2), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),

        ("ALIGN", (0, 1), (0, -2), "CENTER"),
        ("ALIGN", (3, 1), (4, -2), "CENTER"),
        ("ALIGN", (5, 1), (6, -2), "CENTER"),
        ("ALIGN", (7, 1), (7, -2), "RIGHT"),
        ("ALIGN", (9, 1), (9, -2), "RIGHT"),

        ("SPAN", (0, -1), (8, -1)),
        ("ALIGN", (8, -1), (8, -1), "RIGHT"),
        ("ALIGN", (9, -1), (9, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 1.0, colors.black),
    ]))

    elements.append(tbl)
    doc.build(elements)
    return buf.getvalue()



def run_calculation(excel_bytes: bytes, date_from: dt.date, date_to: dt.date) -> Tuple[bytes, bytes]:
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    sheet_list = _find_sheet_name(wb, ["Список"])
    sheet_pay = _find_sheet_name(wb, ["Платежа", "Платежи"])
    sheet_rate = _find_sheet_name(wb, ["Ставка"])

    if not sheet_list or not sheet_rate:
        raise ValueError("Нужны листы: 'Список' и 'Ставка'.")

    df_list = _sheet_to_df(wb, sheet_list)
    df_rate = _sheet_to_df(wb, sheet_rate)

    if df_list.empty:
        raise ValueError("Лист 'Список' пустой.")
    if df_rate.empty:
        raise ValueError("Лист 'Ставка' пустой.")

    # payments optional
    df_pay = pd.DataFrame()
    if sheet_pay:
        df_pay = _sheet_to_df(wb, sheet_pay)

    # required columns in list
    if "Номер договора" not in df_list.columns or "Сумма ОД" not in df_list.columns:
        raise ValueError("В листе 'Список' нужны колонки: 'Номер договора' и 'Сумма ОД'.")

    # optional attrs for PDF
    if "Дата договора" in df_list.columns:
        df_list["Дата договора"] = df_list["Дата договора"].apply(_to_date)
    else:
        df_list["Дата договора"] = None

    if "ФИО" not in df_list.columns:
        df_list["ФИО"] = ""

    rates = _parse_rates(df_rate)

    # build payments dict (optional)
    payments_by: Dict[str, List[Tuple[dt.date, float]]] = {}
    if (
        not df_pay.empty
        and {"Номер договора", "Дата платежа", "Сума платежа"}.issubset(set(df_pay.columns))
    ):
        df_pay["Дата платежа"] = df_pay["Дата платежа"].apply(_to_date)
        for _, r in df_pay.iterrows():
            if pd.isna(r["Номер договора"]):
                continue
            cn = str(int(r["Номер договора"]))
            payments_by.setdefault(cn, []).append((r["Дата платежа"], float(r["Сума платежа"])))
    # else: ignore payments

    totals: Dict[str, float] = {}
    pdfs: Dict[str, bytes] = {}

    for _, r in df_list.iterrows():
        cn = str(int(r["Номер договора"]))
        principal = float(r["Сумма ОД"])
        fio = str(r.get("ФИО", "") or "")
        cdate = r.get("Дата договора", None)

        total, rows = _compute_contract(date_from, date_to, principal, rates, payments_by.get(cn, []))
        totals[cn] = round(total, 2)

        pdfs[cn] = _build_pdf(
            contract_no=cn,
            contract_date=cdate,
            fio=fio,
            date_to=date_to,
            principal=principal,
            rows=rows,
            total=total,
        )

    # write updated excel (preserve sheets)
    wb_out = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb_out[sheet_list]

    header_row = None
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None and str(v).strip() != "" for v in row):
            header_row = i
            headers = [str(v).strip() if v is not None else "" for v in row]
            break
    if header_row is None or headers is None:
        raise ValueError("Не удалось найти заголовок в листе 'Список'.")

    contract_col = headers.index("Номер договора") + 1

    # создаём/находим колонку "Сума по 395"
    if "Сума по 395" in headers:
        percent_col = headers.index("Сума по 395") + 1
    else:
        percent_col = len(headers) + 1
        ws.cell(row=header_row, column=percent_col, value="Сума по 395")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=contract_col).value
        if v is None:
            continue
        cn = str(int(v))
        ws.cell(row=row_idx, column=percent_col, value=totals.get(cn, 0.0))

    out_xlsx_buf = io.BytesIO()
    wb_out.save(out_xlsx_buf)
    out_xlsx_bytes = out_xlsx_buf.getvalue()

    # pack zip
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for cn, pdf_bytes in pdfs.items():
            z.writestr(f"{cn}.pdf", pdf_bytes)

    zip_bytes = zip_buf.getvalue()
    return zip_bytes, out_xlsx_bytes


